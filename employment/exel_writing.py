from openpyxl import load_workbook
import os

def writer(parametr):
    downloads_folder = os.path.join(os.path.expanduser("~"), "Downloads")
    path = os.path.join(downloads_folder, "vacancies.xlsx")

    try:
        # Open existing file
        book = load_workbook(path)
    except FileNotFoundError:
        # If the file doesn't exist — create a new one
        from openpyxl import Workbook
        book = Workbook()

    # Check if a sheet named "товар" exists
    if 'товар' in book.sheetnames:
        sheet = book['товар']
    else:
        sheet = book.create_sheet('товар')

    # Write data to columns A and B
    row = 1
    for item_tit, item_req in zip(parametr['Job Title'], parametr['Requirements']):
        sheet.cell(row=row, column=1, value=item_tit)
        sheet.cell(row=row, column=2, value=item_req)
        row += 1

    # Save changes
    book.save(path)
